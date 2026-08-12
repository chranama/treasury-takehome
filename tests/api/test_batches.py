import asyncio
import csv
import logging
import sqlite3
import time
from io import BytesIO, StringIO
from pathlib import Path
from uuid import UUID

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image

from app.api.request_limits import BATCH_CORRECTION_BODY_BYTES
from app.batches import (
    BATCH_TEMPLATE_HEADERS,
    CONTENT_BEARING_BATCH_TABLES,
    MAX_POLL_RESPONSE_BYTES,
    PreflightIssueCode,
)
from app.batches.drafts import BatchDraftService
from app.config import Settings
from app.db import connect
from app.extraction import (
    ExtractionError,
    ExtractionErrorKind,
    FakeExtractionAdapter,
    FakeExtractionScenario,
    PreparedImage,
)
from app.main import create_app


def make_settings(tmp_path: Path, **overrides: object) -> Settings:
    values: dict[str, object] = {
        "app_env": "test",
        "database_path": tmp_path / "treasury.sqlite3",
        "temp_dir": tmp_path / "tmp",
        "batch_image_dir": tmp_path / "batch-images",
        "frontend_dist_path": tmp_path / "dist",
        "extraction_backend": "fake",
        "live_extraction_enabled": False,
    }
    values.update(overrides)
    return Settings(_env_file=None, **values)


def csv_bytes(rows: list[list[str]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(BATCH_TEMPLATE_HEADERS)
    writer.writerows(rows)
    return output.getvalue().encode()


def png_bytes(color: str = "navy") -> bytes:
    image = Image.new("RGB", (40, 24), color=color)
    output = BytesIO()
    image.save(output, format="PNG")
    image.close()
    return output.getvalue()


def package_files(
    rows: list[list[str]],
    images: list[tuple[str, bytes]],
) -> list[tuple[str, tuple[str, bytes, str]]]:
    files: list[tuple[str, tuple[str, bytes, str]]] = [
        ("spreadsheet", ("batch.csv", csv_bytes(rows), "text/csv"))
    ]
    files.extend(("images", (name, content, "image/png")) for name, content in images)
    return files


def issue_codes(payload: dict) -> set[str]:
    return {issue["code"] for issue in payload["issues"]}


def wait_for_terminal(client: TestClient, batch_id: str) -> dict:
    for _ in range(200):
        payload = client.get(f"/api/batches/{batch_id}").json()
        if payload["state"] in {"completed", "interrupted"}:
            return payload
        time.sleep(0.01)
    raise AssertionError("batch did not reach a terminal state")


class TrackingAdapter:
    def __init__(
        self,
        *,
        fail_calls: set[int] | None = None,
        delay: float = 0,
        call_delays: dict[int, float] | None = None,
        scenarios: dict[int, FakeExtractionScenario] | None = None,
    ) -> None:
        self.delegate = FakeExtractionAdapter()
        self.fail_calls = fail_calls or set()
        self.delay = delay
        self.call_delays = call_delays or {}
        self.scenarios = scenarios or {}
        self.calls = 0
        self.active = 0
        self.max_active = 0

    async def extract(self, image: PreparedImage):
        self.calls += 1
        call = self.calls
        self.active += 1
        self.max_active = max(self.max_active, self.active)
        try:
            delay = self.call_delays.get(call, self.delay)
            if delay:
                await asyncio.sleep(delay)
            if call in self.fail_calls:
                raise ExtractionError(
                    kind=ExtractionErrorKind.UNAVAILABLE,
                    safe_message="provider unavailable",
                    retryable=False,
                )
            scenario = self.scenarios.get(call)
            if scenario is not None:
                return await FakeExtractionAdapter(scenario=scenario).extract(image)
            return await self.delegate.extract(image)
        finally:
            self.active -= 1


def test_downloadable_templates_have_stable_names_and_headers(tmp_path: Path) -> None:
    with TestClient(create_app(make_settings(tmp_path))) as client:
        csv_response = client.get("/api/batch-template.csv")
        xlsx_response = client.get("/api/batch-template.xlsx")

    assert csv_response.status_code == 200
    assert csv_response.headers["content-disposition"] == (
        'attachment; filename="label-review-batch.csv"'
    )
    assert next(csv.reader(StringIO(csv_response.text))) == list(BATCH_TEMPLATE_HEADERS)
    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["content-disposition"] == (
        'attachment; filename="label-review-batch.xlsx"'
    )
    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
    assert tuple(cell.value for cell in next(workbook["Batch"].iter_rows())) == (
        BATCH_TEMPLATE_HEADERS
    )
    workbook.close()


def test_valid_preflight_returns_a_refreshable_draft_and_case_detail(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    rows = [["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]]
    with TestClient(create_app(settings)) as client:
        response = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes())]),
        )
        payload = response.json()
        batch_id = payload["batch_id"]
        case_id = payload["cases"][0]["case_id"]
        recovered = client.get(f"/api/batches/{batch_id}")
        detail = client.get(f"/api/batches/{batch_id}/cases/{case_id}")

    assert response.status_code == 201
    assert response.headers["location"] == f"/api/batches/{batch_id}"
    assert UUID(batch_id).version == 4
    assert payload["counts"]["ready"] == 1
    assert payload["counts"]["needs_correction"] == 0
    assert recovered.status_code == 200
    assert recovered.json() == payload
    assert detail.status_code == 200
    assert detail.json()["expected_input"] == {
        "brand_name": "Brand",
        "class_type": "Bourbon",
        "expected_abv": "45",
        "expected_net_contents": "750 mL",
    }
    assert detail.json()["normalized_expected"] is not None


def test_mixed_draft_correction_and_replacement_update_only_the_target_case(
    tmp_path: Path,
) -> None:
    settings = make_settings(tmp_path)
    rows = [
        ["APP-1", "one.png", "Brand", "Bourbon", "45", "750 mL"],
        ["APP-2", "missing.png", "Brand", "Bourbon", "101", "750 mL"],
    ]
    with TestClient(create_app(settings)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("one.png", png_bytes("red"))]),
        ).json()
        batch_id = created["batch_id"]
        ready_case, correction_case = created["cases"]
        ready_before = client.get(f"/api/batches/{batch_id}/cases/{ready_case['case_id']}").json()

        corrected = client.patch(
            f"/api/batches/{batch_id}/cases/{correction_case['case_id']}",
            json={"expected_abv": "45%"},
        )
        replaced = client.put(
            f"/api/batches/{batch_id}/cases/{correction_case['case_id']}/image",
            files={"image": ("replacement.png", png_bytes("blue"), "image/png")},
        )
        refreshed = client.get(f"/api/batches/{batch_id}").json()
        ready_after = client.get(f"/api/batches/{batch_id}/cases/{ready_case['case_id']}").json()

    assert created["counts"] == {
        "total": 2,
        "needs_correction": 1,
        "ready": 1,
        "queued": 0,
        "processing": 0,
        "completed": 0,
        "failed": 0,
        "interrupted": 0,
        "not_selected": 0,
    }
    assert corrected.status_code == 200
    assert corrected.json()["summary"]["state"] == "needs_correction"
    assert issue_codes(corrected.json()["summary"]) == {PreflightIssueCode.MISSING_IMAGE.value}
    assert replaced.status_code == 200
    assert replaced.json()["summary"]["state"] == "ready"
    assert replaced.json()["summary"]["label_image_filename"] == "replacement.png"
    assert refreshed["counts"]["ready"] == 2
    assert ready_after == ready_before


def test_structurally_invalid_and_entirely_invalid_packages_return_safe_issues(
    tmp_path: Path,
) -> None:
    settings = make_settings(tmp_path)
    entirely_invalid_rows = [["", "missing.png", "", "", "101", "750 oz"]]
    with TestClient(create_app(settings)) as client:
        malformed = client.post(
            "/api/batches/preflight",
            files={"spreadsheet": ("batch.csv", b'"unterminated', "text/csv")},
        )
        entirely_invalid = client.post(
            "/api/batches/preflight",
            files=package_files(entirely_invalid_rows, []),
        )

    assert malformed.status_code == 422
    malformed_payload = malformed.json()
    assert issue_codes(malformed_payload) == {PreflightIssueCode.MALFORMED_SPREADSHEET.value}
    assert UUID(malformed_payload["correlation_id"]).version == 4
    assert entirely_invalid.status_code == 201
    invalid_payload = entirely_invalid.json()
    assert invalid_payload["counts"]["ready"] == 0
    assert invalid_payload["counts"]["needs_correction"] == 1
    assert {
        PreflightIssueCode.MISSING_APPLICATION_ID.value,
        PreflightIssueCode.MISSING_IMAGE.value,
        PreflightIssueCode.INVALID_BRAND.value,
        PreflightIssueCode.INVALID_CLASS_TYPE.value,
        PreflightIssueCode.INVALID_ABV.value,
        PreflightIssueCode.INVALID_NET_CONTENTS.value,
    } <= issue_codes(invalid_payload["cases"][0])


def test_missing_multipart_fields_return_batch_specific_safe_issues(tmp_path: Path) -> None:
    with TestClient(create_app(make_settings(tmp_path))) as client:
        missing_spreadsheet = client.post("/api/batches/preflight")

    assert missing_spreadsheet.status_code == 422
    assert issue_codes(missing_spreadsheet.json()) == {
        PreflightIssueCode.UNSUPPORTED_SPREADSHEET.value
    }


def test_invalid_replacement_preserves_the_draft_and_returns_row_issue(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    rows = [["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]]
    with TestClient(create_app(settings)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes())]),
        ).json()
        batch_id = created["batch_id"]
        case_id = created["cases"][0]["case_id"]
        rejected = client.put(
            f"/api/batches/{batch_id}/cases/{case_id}/image",
            files={"image": ("broken.png", b"not an image", "image/png")},
        )
        recovered = client.get(f"/api/batches/{batch_id}")

    assert rejected.status_code == 422
    assert issue_codes(rejected.json()) == {PreflightIssueCode.UNSUPPORTED_IMAGE.value}
    assert recovered.status_code == 200
    assert recovered.json()["counts"]["ready"] == 1
    assert len(list(settings.batch_image_dir.iterdir())) == 1
    assert list(settings.temp_dir.iterdir()) == []


def test_unknown_malformed_and_cross_batch_ids_share_the_not_found_response(
    tmp_path: Path,
) -> None:
    settings = make_settings(tmp_path)
    rows = [["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]]
    with TestClient(create_app(settings)) as client:
        first = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes("red"))]),
        ).json()
        second = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes("blue"))]),
        ).json()
        responses = [
            client.get("/api/batches/not-a-uuid"),
            client.get(f"/api/batches/{first['batch_id']}/cases/not-a-uuid"),
            client.get(f"/api/batches/{first['batch_id']}/cases/{second['cases'][0]['case_id']}"),
        ]
        collection = client.get("/api/batches")
        unknown_export = client.get("/api/batches/not-a-uuid/results.csv")

    assert collection.status_code == 404
    for response in [*responses, unknown_export]:
        assert response.status_code == 404
        payload = response.json()
        assert payload["code"] == "batch_not_found"
        assert payload["message"] == "The requested batch is unavailable."
        assert UUID(payload["correlation_id"]).version == 4


def test_draft_export_conflicts_and_expired_poll_detail_and_export_are_not_found(
    tmp_path: Path,
) -> None:
    settings = make_settings(tmp_path)
    rows = [["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]]
    with TestClient(create_app(settings)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes())]),
        ).json()
        batch_id = created["batch_id"]
        case_id = created["cases"][0]["case_id"]
        draft_export = client.get(f"/api/batches/{batch_id}/results.csv")
        with connect(settings.database_path) as connection:
            connection.execute(
                "UPDATE batch_reviews SET expires_at = '2000-01-01T00:00:00+00:00' "
                "WHERE batch_id = ?",
                (batch_id,),
            )
        expired = [
            client.get(f"/api/batches/{batch_id}"),
            client.get(f"/api/batches/{batch_id}/cases/{case_id}"),
            client.get(f"/api/batches/{batch_id}/results.csv"),
        ]

    assert draft_export.status_code == 409
    assert draft_export.json()["code"] == "batch_results_unavailable"
    for response in expired:
        assert response.status_code == 404
        assert response.json()["code"] == "batch_not_found"


def test_maximum_case_polling_representation_stays_within_the_public_byte_bound(
    tmp_path: Path,
) -> None:
    rows = [
        [f"APP-{index:02d}", f"label-{index:02d}.png", "Brand", "Bourbon", "45", "750 mL"]
        for index in range(1, 26)
    ]
    images = [(f"label-{index:02d}.png", png_bytes()) for index in range(1, 26)]
    with TestClient(create_app(make_settings(tmp_path))) as client:
        created = client.post("/api/batches/preflight", files=package_files(rows, images))
        polled = client.get(f"/api/batches/{created.json()['batch_id']}")

    assert created.status_code == 201
    assert polled.status_code == 200
    assert polled.headers["cache-control"] == "no-store"
    assert len(polled.content) <= MAX_POLL_RESPONSE_BYTES
    assert polled.json()["counts"]["total"] == 25


def test_correction_and_start_json_requests_have_route_specific_body_bounds(
    tmp_path: Path,
) -> None:
    rows = [["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]]
    oversized = b"{" + b"x" * BATCH_CORRECTION_BODY_BYTES
    with TestClient(create_app(make_settings(tmp_path))) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes())]),
        ).json()
        case_id = created["cases"][0]["case_id"]
        correction = client.patch(
            f"/api/batches/{created['batch_id']}/cases/{case_id}",
            content=oversized,
            headers={"Content-Type": "application/json"},
        )
        start = client.post(
            f"/api/batches/{created['batch_id']}/start",
            content=oversized,
            headers={
                "Content-Type": "application/json",
                "Idempotency-Key": "bounded-start-request",
            },
        )

    for response in (correction, start):
        assert response.status_code == 413
        assert response.json()["code"] == "batch_request_too_large"
        assert response.json()["message"] == "The batch request exceeds the allowed size."


def test_twenty_five_case_fake_batch_keeps_concurrency_two_and_isolates_outcomes(
    tmp_path: Path,
) -> None:
    rows = [
        [
            f"APP-{index:02d}",
            f"label-{index:02d}.png",
            "Treasury Reserve",
            "Kentucky Straight Bourbon Whiskey",
            "45",
            "750 mL",
        ]
        for index in range(1, 26)
    ]
    images = [(f"label-{index:02d}.png", png_bytes()) for index in range(1, 26)]
    adapter = TrackingAdapter(
        fail_calls={5, 18},
        delay=0.002,
        scenarios={
            2: FakeExtractionScenario.MISMATCHED_NET_CONTENTS,
            9: FakeExtractionScenario.UNREADABLE_IMAGE,
            14: FakeExtractionScenario.ALTERED_WARNING_TEXT,
            23: FakeExtractionScenario.BRAND_MISMATCH,
        },
    )

    with TestClient(create_app(make_settings(tmp_path), extraction_adapter=adapter)) as client:
        created = client.post("/api/batches/preflight", files=package_files(rows, images)).json()
        started = client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "twenty-five-case-fake-batch"},
        )
        completed = wait_for_terminal(client, created["batch_id"])

    assert started.status_code == 202
    assert completed["state"] == "completed"
    assert completed["counts"]["completed"] == 23
    assert completed["counts"]["failed"] == 2
    assert adapter.calls == 25
    assert adapter.max_active == 2
    assert {case["outcome"] for case in completed["cases"] if case["state"] == "completed"} == {
        "all_checks_passed",
        "needs_review",
    }
    assert completed["cases"][-1]["state"] == "completed"
    assert not list(make_settings(tmp_path).batch_image_dir.glob("*.png"))


def test_processed_image_is_deleted_while_later_cases_remain_active(tmp_path: Path) -> None:
    rows = [
        [f"APP-{index}", f"label-{index}.png", "Brand", "Bourbon", "45", "750 mL"]
        for index in range(1, 4)
    ]
    images = [(f"label-{index}.png", png_bytes()) for index in range(1, 4)]
    adapter = TrackingAdapter(call_delays={2: 0.3, 3: 0.3})
    settings = make_settings(tmp_path)

    with TestClient(create_app(settings, extraction_adapter=adapter)) as client:
        created = client.post("/api/batches/preflight", files=package_files(rows, images)).json()
        client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "immediate-image-deletion"},
        )
        active = None
        for _ in range(100):
            candidate = client.get(f"/api/batches/{created['batch_id']}").json()
            if candidate["counts"]["completed"] == 1 and (
                candidate["counts"]["queued"] + candidate["counts"]["processing"] > 0
            ) and len(list(settings.batch_image_dir.glob("*.png"))) == 2:
                active = candidate
                break
            time.sleep(0.005)
        assert active is not None
        assert len(list(settings.batch_image_dir.glob("*.png"))) == 2
        terminal = wait_for_terminal(client, created["batch_id"])

    assert terminal["counts"]["completed"] == 3
    assert not list(settings.batch_image_dir.glob("*.png"))


def test_expiry_removes_unselected_image_and_every_content_bearing_batch_row(
    tmp_path: Path,
) -> None:
    settings = make_settings(tmp_path)
    rows = [
        ["APP-1", "one.png", "Brand", "Bourbon", "45", "750 mL"],
        ["APP-2", "two.png", "Brand", "Bourbon", "101", "750 mL"],
    ]
    with TestClient(create_app(settings)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(
                rows,
                [("one.png", png_bytes("red")), ("two.png", png_bytes("blue"))],
            ),
        ).json()
        client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "ready_cases_only"},
            headers={"Idempotency-Key": "expiry-ready-only-batch"},
        )
        completed = wait_for_terminal(client, created["batch_id"])

    assert completed["counts"]["completed"] == 1
    assert completed["counts"]["not_selected"] == 1
    assert len(list(settings.batch_image_dir.glob("*.png"))) == 1
    with connect(settings.database_path) as connection:
        before = {
            table: connection.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            for table in CONTENT_BEARING_BATCH_TABLES
        }
        connection.execute(
            "UPDATE batch_reviews SET expires_at = '2000-01-01T00:00:00+00:00' "
            "WHERE batch_id = ?",
            (created["batch_id"],),
        )
    assert before == {
        "batch_reviews": 1,
        "batch_images": 2,
        "batch_cases": 2,
        "batch_case_results": 1,
    }

    cleanup = BatchDraftService(
        database_path=settings.database_path,
        image_dir=settings.batch_image_dir,
        temp_dir=settings.temp_dir,
    )
    result = asyncio.run(cleanup.cleanup_expired_and_orphaned())

    assert result.expired_batch_count == 1
    assert result.deleted_file_count == 1
    assert not list(settings.batch_image_dir.glob("*.png"))
    with connect(settings.database_path) as connection:
        assert all(
            connection.execute(f"SELECT COUNT(*) FROM {table}").fetchone() == (0,)
            for table in CONTENT_BEARING_BATCH_TABLES
        )


def test_logs_and_operational_rows_exclude_batch_content_and_source_address(
    tmp_path: Path,
    caplog,
) -> None:
    caplog.set_level(logging.DEBUG)
    settings = make_settings(
        tmp_path,
        extraction_backend="openai",
        live_extraction_enabled=True,
        openai_api_key="test-key",
        openai_transient_retries=0,
        live_daily_attempt_limit=10,
        live_cumulative_cost_limit_usd="1",
        live_attempt_reservation_usd="0.01",
        live_source_window_seconds=60,
        live_source_max_submissions=10,
        trust_cloudflare_client_ip=True,
    )
    forbidden = {
        "PRIVATE-APPLICATION-ROW",
        "private-workbook-filename.png",
        "EXPECTED PRIVATE BRAND",
        "EXPECTED PRIVATE CLASS",
        "Treasury Reserve",
        "Kentucky Straight Bourbon Whiskey",
        "198.51.100.44",
    }
    rows = [
        [
            "PRIVATE-APPLICATION-ROW",
            "private-workbook-filename.png",
            "EXPECTED PRIVATE BRAND",
            "EXPECTED PRIVATE CLASS",
            "45",
            "750 mL",
        ]
    ]
    headers = {"CF-Connecting-IP": "198.51.100.44"}

    with TestClient(create_app(settings, extraction_adapter=TrackingAdapter())) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("private-workbook-filename.png", png_bytes())]),
            headers=headers,
        ).json()
        client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "all_cases"},
            headers={**headers, "Idempotency-Key": "content-free-ledger-test"},
        )
        completed = wait_for_terminal(client, created["batch_id"])

    assert completed["counts"]["completed"] == 1
    with connect(settings.database_path) as connection:
        connection.row_factory = sqlite3.Row
        attempt_count = connection.execute("SELECT COUNT(*) FROM provider_attempts").fetchone()[0]
        operational = repr(
            {
                "review_submissions": [
                    dict(row) for row in connection.execute("SELECT * FROM review_submissions")
                ],
                "provider_attempts": [
                    dict(row) for row in connection.execute("SELECT * FROM provider_attempts")
                ],
            }
        )
        assert attempt_count == 1

    for value in forbidden:
        assert value not in caplog.text
        assert value not in operational


def test_start_is_idempotent_and_each_internal_case_has_one_durable_attempt(
    tmp_path: Path,
) -> None:
    settings = make_settings(
        tmp_path,
        extraction_backend="openai",
        live_extraction_enabled=True,
        openai_api_key="test-key",
        openai_transient_retries=0,
        live_daily_attempt_limit=10,
        live_cumulative_cost_limit_usd="1",
        live_attempt_reservation_usd="0.01",
        live_source_window_seconds=60,
        live_source_max_submissions=2,
    )
    adapter = TrackingAdapter(delay=0.02)
    rows = [
        [f"APP-{index}", f"label-{index}.png", "Brand", "Bourbon", "45", "750 mL"]
        for index in range(1, 4)
    ]
    images = [(f"label-{index}.png", png_bytes()) for index in range(1, 4)]

    with TestClient(create_app(settings, extraction_adapter=adapter)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, images),
        ).json()
        path = f"/api/batches/{created['batch_id']}/start"
        headers = {"Idempotency-Key": "one-stable-batch-key"}
        first = client.post(path, json={"selection": "all_cases"}, headers=headers)
        repeated = client.post(path, json={"selection": "all_cases"}, headers=headers)
        completed = wait_for_terminal(client, created["batch_id"])
        repeated_after_completion = client.post(
            path,
            json={"selection": "all_cases"},
            headers=headers,
        )
        conflicting = client.post(
            path,
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "a-different-batch-key"},
        )

    assert first.status_code == 202
    assert repeated.status_code == 202
    assert repeated_after_completion.status_code == 202
    assert conflicting.status_code == 409
    assert conflicting.json()["code"] == "batch_state_conflict"
    assert completed["counts"]["completed"] == 3
    assert completed["counts"]["failed"] == 0
    assert adapter.calls == 3
    assert adapter.max_active == 2
    with connect(settings.database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM review_submissions").fetchone() == (3,)
        assert connection.execute("SELECT COUNT(*) FROM provider_attempts").fetchone() == (3,)


def test_ready_only_selection_is_explicit_and_case_failure_is_isolated(tmp_path: Path) -> None:
    adapter = TrackingAdapter(fail_calls={1}, delay=0.01)
    rows = [
        ["APP-1", "one.png", "Brand", "Bourbon", "45", "750 mL"],
        ["APP-2", "two.png", "Brand", "Bourbon", "45", "750 mL"],
        ["APP-3", "missing.png", "Brand", "Bourbon", "101", "750 mL"],
    ]
    with TestClient(create_app(make_settings(tmp_path), extraction_adapter=adapter)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(
                rows,
                [("one.png", png_bytes("red")), ("two.png", png_bytes("blue"))],
            ),
        ).json()
        path = f"/api/batches/{created['batch_id']}/start"
        rejected = client.post(
            path,
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "all-cases-with-errors"},
        )
        started = client.post(
            path,
            json={"selection": "ready_cases_only"},
            headers={"Idempotency-Key": "ready-cases-only-key"},
        )
        completed = wait_for_terminal(client, created["batch_id"])
        completed_case = next(case for case in completed["cases"] if case["state"] == "completed")
        failed_case = next(case for case in completed["cases"] if case["state"] == "failed")
        completed_detail = client.get(
            f"/api/batches/{created['batch_id']}/cases/{completed_case['case_id']}"
        )
        failed_detail = client.get(
            f"/api/batches/{created['batch_id']}/cases/{failed_case['case_id']}"
        )
        exported = client.get(f"/api/batches/{created['batch_id']}/results.csv")

    assert rejected.status_code == 409
    assert rejected.json()["code"] == "batch_has_corrections"
    assert started.status_code == 202
    assert completed["state"] == "completed"
    assert completed["counts"]["completed"] == 1
    assert completed["counts"]["failed"] == 1
    assert completed["counts"]["not_selected"] == 1
    assert completed["expires_at"] == created["expires_at"]
    assert "provider_request_id" not in str(completed)
    assert "result_json" not in str(completed)
    assert completed_case["short_reason"] == (
        "Visible brand name differs materially from the expected value."
    )
    assert failed_case["short_reason"] == "Label extraction is temporarily unavailable. Try again."
    assert completed_detail.status_code == 200
    assert completed_detail.headers["cache-control"] == "no-store"
    detail_checks = completed_detail.json()["result"]["result"]["checks"]
    assert {check["name"] for check in detail_checks} == {
        "brand_name",
        "class_type",
        "alcohol_content",
        "net_contents",
        "government_warning",
    }
    assert failed_detail.status_code == 200
    assert failed_detail.json()["result"] is None
    assert exported.status_code == 200
    assert exported.headers["content-disposition"] == (
        'attachment; filename="label-review-results.csv"'
    )
    assert exported.headers["cache-control"] == "no-store"
    export_rows = list(csv.DictReader(StringIO(exported.content.decode("utf-8-sig"))))
    assert len(export_rows) == 2
    assert {row["Processing Status"] for row in export_rows} == {"completed", "failed"}
    assert {row["Application ID"] for row in export_rows} == {"APP-1", "APP-2"}
    completed_row = next(row for row in export_rows if row["Processing Status"] == "completed")
    assert completed_row["Brand Name Status"] == "needs_review"
    assert completed_row["Government Warning Status"] == "match"
    assert completed_row["Short Reason"] == completed_case["short_reason"]
    assert adapter.calls == 2
    assert not list(make_settings(tmp_path).batch_image_dir.glob("*.png"))


def test_capacity_exhaustion_creates_per_case_failures_without_unaccounted_requests(
    tmp_path: Path,
) -> None:
    settings = make_settings(
        tmp_path,
        extraction_backend="openai",
        live_extraction_enabled=True,
        openai_api_key="test-key",
        openai_transient_retries=0,
        live_daily_attempt_limit=1,
        live_cumulative_cost_limit_usd="1",
        live_attempt_reservation_usd="0.01",
        live_source_window_seconds=60,
        live_source_max_submissions=10,
    )
    adapter = TrackingAdapter(delay=0.01)
    rows = [
        [f"APP-{index}", f"label-{index}.png", "Brand", "Bourbon", "45", "750 mL"]
        for index in range(1, 4)
    ]
    images = [(f"label-{index}.png", png_bytes()) for index in range(1, 4)]

    with TestClient(create_app(settings, extraction_adapter=adapter)) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, images),
        ).json()
        started = client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "capacity-batch-key"},
        )
        completed = wait_for_terminal(client, created["batch_id"])

    assert started.status_code == 202
    assert completed["counts"]["completed"] == 1
    assert completed["counts"]["failed"] == 2
    assert {case["short_reason"] for case in completed["cases"] if case["state"] == "failed"} == {
        "Live review capacity is temporarily unavailable."
    }
    assert adapter.calls == 1
    with connect(settings.database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM provider_attempts").fetchone() == (1,)
        assert connection.execute(
            "SELECT COUNT(*) FROM review_submissions "
            "WHERE status = 'failed' AND error_kind = 'capacity_reached'"
        ).fetchone() == (2,)


def test_startup_marks_uncertain_batch_work_interrupted_without_replay(tmp_path: Path) -> None:
    settings = make_settings(
        tmp_path,
        extraction_backend="openai",
        live_extraction_enabled=True,
        openai_api_key="test-key",
        openai_transient_retries=0,
        live_daily_attempt_limit=10,
        live_cumulative_cost_limit_usd="1",
        live_attempt_reservation_usd="0.01",
        live_source_window_seconds=60,
        live_source_max_submissions=10,
    )
    rows = [["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]]
    with TestClient(create_app(settings, extraction_adapter=TrackingAdapter())) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(rows, [("label.png", png_bytes())]),
        ).json()

    case_id = created["cases"][0]["case_id"]
    with connect(settings.database_path) as connection:
        connection.execute(
            "UPDATE batch_reviews SET status = 'processing' WHERE batch_id = ?",
            (created["batch_id"],),
        )
        connection.execute(
            """
            INSERT INTO review_submissions (
                idempotency_hash, correlation_id, status, created_at
            ) VALUES ('uncertain-hash', ?, 'processing', '2026-08-12T12:00:00+00:00')
            """,
            (case_id,),
        )
        connection.execute(
            """
            INSERT INTO provider_attempts (
                correlation_id, attempt_number, status, reserved_at,
                reserved_cost_units, model, prompt_revision, image_detail,
                requested_service_tier
            ) VALUES (?, 1, 'reserved', '2026-08-12T12:00:00+00:00',
                      1000000, 'gpt-5.6-luna', 'label-observations-v2', 'high', 'default')
            """,
            (case_id,),
        )
        connection.execute(
            "UPDATE batch_cases SET status = 'processing', provider_correlation_id = ? "
            "WHERE case_id = ?",
            (case_id, case_id),
        )
        connection.execute(
            "UPDATE batch_images SET status = 'processing' WHERE batch_id = ?",
            (created["batch_id"],),
        )

    adapter = TrackingAdapter()
    with TestClient(create_app(settings, extraction_adapter=adapter)) as client:
        reconciled = client.get(f"/api/batches/{created['batch_id']}").json()

    assert reconciled["state"] == "interrupted"
    assert reconciled["counts"]["interrupted"] == 1
    assert adapter.calls == 0
    assert not list(settings.batch_image_dir.glob("*.png"))
    with connect(settings.database_path) as connection:
        assert connection.execute(
            "SELECT status, error_kind FROM review_submissions WHERE correlation_id = ?",
            (case_id,),
        ).fetchone() == ("failed", "interrupted")
        assert connection.execute(
            "SELECT status, error_kind FROM provider_attempts WHERE correlation_id = ?",
            (case_id,),
        ).fetchone() == ("failed", "interrupted")


def test_mid_batch_restart_preserves_completed_case_and_never_replays_uncertain_attempts(
    tmp_path: Path,
) -> None:
    settings = make_settings(
        tmp_path,
        extraction_backend="openai",
        live_extraction_enabled=True,
        openai_api_key="test-key",
        openai_transient_retries=0,
        live_daily_attempt_limit=10,
        live_cumulative_cost_limit_usd="1",
        live_attempt_reservation_usd="0.01",
        live_source_window_seconds=60,
        live_source_max_submissions=10,
    )
    rows = [
        [f"APP-{index}", f"label-{index}.png", "Brand", "Bourbon", "45", "750 mL"]
        for index in range(1, 4)
    ]
    images = [(f"label-{index}.png", png_bytes()) for index in range(1, 4)]
    first_adapter = TrackingAdapter(delay=1, call_delays={1: 0})
    application = create_app(settings, extraction_adapter=first_adapter)
    application.state.batch_processing_service.drain_seconds = 0.01

    with TestClient(application) as client:
        created = client.post("/api/batches/preflight", files=package_files(rows, images)).json()
        client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "mid-batch-restart-key"},
        )
        for _ in range(100):
            active = client.get(f"/api/batches/{created['batch_id']}").json()
            if active["counts"]["completed"] == 1 and active["counts"]["processing"] == 2:
                break
            time.sleep(0.005)
        else:
            raise AssertionError("batch did not reach the intended mid-batch state")

    with connect(settings.database_path) as connection:
        attempt_count_before_restart = connection.execute(
            "SELECT COUNT(*) FROM provider_attempts"
        ).fetchone()[0]
    assert attempt_count_before_restart == 3

    restarted_adapter = TrackingAdapter()
    with TestClient(create_app(settings, extraction_adapter=restarted_adapter)) as client:
        recovered = client.get(f"/api/batches/{created['batch_id']}").json()

    assert recovered["state"] == "interrupted"
    assert recovered["counts"]["completed"] == 1
    assert recovered["counts"]["interrupted"] == 2
    assert restarted_adapter.calls == 0
    with connect(settings.database_path) as connection:
        attempt_states = connection.execute(
            "SELECT status, error_kind FROM provider_attempts ORDER BY id"
        ).fetchall()
        assert connection.execute("SELECT COUNT(*) FROM provider_attempts").fetchone() == (3,)
        assert connection.execute(
            "SELECT COUNT(*) FROM provider_attempts WHERE status = 'reserved'"
        ).fetchone() == (0,)
        assert connection.execute(
            "SELECT COUNT(*) FROM provider_attempts "
            "WHERE status = 'failed' AND error_kind != 'interrupted'"
        ).fetchone() == (0,), attempt_states
        assert connection.execute(
            "SELECT COUNT(*) FROM review_submissions WHERE status = 'completed'"
        ).fetchone() == (1,)
        assert connection.execute(
            "SELECT COUNT(*) FROM review_submissions "
            "WHERE status = 'failed' AND error_kind = 'interrupted'"
        ).fetchone() == (2,)


def test_graceful_shutdown_bounds_drain_and_interrupts_remaining_cases(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    adapter = TrackingAdapter(delay=1)
    application = create_app(settings, extraction_adapter=adapter)
    application.state.batch_processing_service.drain_seconds = 0.01
    rows = [
        ["APP-1", "one.png", "Brand", "Bourbon", "45", "750 mL"],
        ["APP-2", "two.png", "Brand", "Bourbon", "45", "750 mL"],
    ]

    with TestClient(application) as client:
        created = client.post(
            "/api/batches/preflight",
            files=package_files(
                rows,
                [("one.png", png_bytes("red")), ("two.png", png_bytes("blue"))],
            ),
        ).json()
        client.post(
            f"/api/batches/{created['batch_id']}/start",
            json={"selection": "all_cases"},
            headers={"Idempotency-Key": "shutdown-drain-key"},
        )
        for _ in range(100):
            active = client.get(f"/api/batches/{created['batch_id']}").json()
            if active["counts"]["processing"]:
                break
            time.sleep(0.005)

    with connect(settings.database_path) as connection:
        assert connection.execute(
            "SELECT status FROM batch_reviews WHERE batch_id = ?",
            (created["batch_id"],),
        ).fetchone() == ("interrupted",)
        assert connection.execute(
            "SELECT COUNT(*) FROM batch_cases WHERE batch_id = ? AND status = 'interrupted'",
            (created["batch_id"],),
        ).fetchone() == (2,)
    assert not list(settings.batch_image_dir.glob("*.png"))
