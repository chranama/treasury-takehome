import csv
from io import BytesIO, StringIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from app.batches import (
    BATCH_TEMPLATE_HEADERS,
    BatchField,
    PreflightIssueCode,
    SpreadsheetKind,
    generate_xlsx_template,
    is_base_filename,
    normalize_filename,
    parse_spreadsheet,
)


def csv_package(rows: list[list[object]], *, bom: bool = False) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(BATCH_TEMPLATE_HEADERS)
    writer.writerows(rows)
    encoded = output.getvalue().encode("utf-8")
    return (b"\xef\xbb\xbf" + encoded) if bom else encoded


def xlsx_package(rows: list[list[object]]) -> bytes:
    workbook = load_workbook(BytesIO(generate_xlsx_template()))
    batch = workbook["Batch"]
    for row in rows:
        batch.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def parse_bytes(tmp_path: Path, content: bytes, *, filename: str):
    path = tmp_path / filename
    path.write_bytes(content)
    return parse_spreadsheet(path, filename=filename)


def issue_codes(parsed) -> list[PreflightIssueCode]:
    return [issue.code for issue in parsed.issues]


def row_issue_codes(parsed, index: int = 0) -> list[PreflightIssueCode]:
    return [issue.code for issue in parsed.rows[index].issues]


@pytest.mark.parametrize("kind", [SpreadsheetKind.CSV, SpreadsheetKind.XLSX])
def test_valid_csv_and_xlsx_rows_normalize_to_identical_contracts(
    tmp_path: Path,
    kind: SpreadsheetKind,
) -> None:
    row = [
        " APP-001 ",
        " Label.PNG ",
        " OLD TOM DISTILLERY ",
        " Kentucky Straight Bourbon Whiskey ",
        "45%",
        "0.75 L",
    ]
    content = csv_package([row], bom=True) if kind == SpreadsheetKind.CSV else xlsx_package([row])
    parsed = parse_bytes(tmp_path, content, filename=f"batch.{kind.value}")

    assert parsed.kind == kind
    assert parsed.issues == ()
    assert len(parsed.rows) == 1
    result = parsed.rows[0]
    assert result.application_id == "APP-001"
    assert result.normalized_application_id == "app-001"
    assert result.label_image_filename == "Label.PNG"
    assert result.normalized_label_image_filename == "label.png"
    assert result.expected_input.model_dump() == {
        "brand_name": "OLD TOM DISTILLERY",
        "class_type": "Kentucky Straight Bourbon Whiskey",
        "expected_abv": "45%",
        "expected_net_contents": "0.75 L",
    }
    assert result.normalized_expected is not None
    assert str(result.normalized_expected.abv) == "45"
    assert str(result.normalized_expected.net_contents.value) == "0.75"
    assert result.normalized_expected.net_contents.unit.value == "L"
    assert result.issues == ()


@pytest.mark.parametrize(
    ("headers", "expected_code"),
    [
        (list(BATCH_TEMPLATE_HEADERS[:-1]), PreflightIssueCode.MISSING_REQUIRED_COLUMN),
        (
            [*BATCH_TEMPLATE_HEADERS[:-1], "Unexpected"],
            PreflightIssueCode.UNEXPECTED_COLUMN,
        ),
        (
            [BATCH_TEMPLATE_HEADERS[0], BATCH_TEMPLATE_HEADERS[0], *BATCH_TEMPLATE_HEADERS[2:]],
            PreflightIssueCode.DUPLICATE_COLUMN,
        ),
        (
            [BATCH_TEMPLATE_HEADERS[1], BATCH_TEMPLATE_HEADERS[0], *BATCH_TEMPLATE_HEADERS[2:]],
            PreflightIssueCode.INVALID_COLUMN_ORDER,
        ),
    ],
)
def test_csv_headers_are_exact_and_ordered(
    tmp_path: Path,
    headers: list[str],
    expected_code: PreflightIssueCode,
) -> None:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"])

    parsed = parse_bytes(tmp_path, output.getvalue().encode(), filename="batch.csv")

    assert expected_code in issue_codes(parsed)
    assert parsed.rows == ()


@pytest.mark.parametrize(
    ("content", "expected_code"),
    [
        (b"\xff\xfeinvalid", PreflightIssueCode.INVALID_CSV_ENCODING),
        (b"Application ID\x00,Label Image Filename", PreflightIssueCode.NUL_BYTE_NOT_ALLOWED),
        (b'"unterminated', PreflightIssueCode.MALFORMED_SPREADSHEET),
    ],
)
def test_csv_rejects_invalid_encoding_nul_and_malformed_records(
    tmp_path: Path,
    content: bytes,
    expected_code: PreflightIssueCode,
) -> None:
    parsed = parse_bytes(tmp_path, content, filename="batch.csv")

    assert issue_codes(parsed) == [expected_code]


def test_empty_and_over_limit_batches_are_explicit(tmp_path: Path) -> None:
    empty = parse_bytes(tmp_path, csv_package([]), filename="empty.csv")
    rows = [
        [f"APP-{index}", f"label-{index}.png", "Brand", "Bourbon", "45", "750 mL"]
        for index in range(26)
    ]
    over_limit = parse_bytes(tmp_path, csv_package(rows), filename="over.csv")

    assert issue_codes(empty) == [PreflightIssueCode.EMPTY_BATCH]
    assert issue_codes(over_limit) == [PreflightIssueCode.TOO_MANY_CASES]
    assert len(over_limit.rows) == 25


def test_duplicate_ids_and_image_references_mark_every_affected_row(tmp_path: Path) -> None:
    rows = [
        ["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"],
        [" app-1 ", "LABEL.PNG", "Brand", "Bourbon", "45", "750 mL"],
    ]
    parsed = parse_bytes(tmp_path, csv_package(rows), filename="duplicates.csv")

    for row in parsed.rows:
        codes = [issue.code for issue in row.issues]
        assert PreflightIssueCode.DUPLICATE_APPLICATION_ID in codes
        assert PreflightIssueCode.DUPLICATE_IMAGE_FILENAME in codes


@pytest.mark.parametrize(
    "filename",
    ["../label.png", "folder/label.png", r"folder\label.png", r"C:\label.png", ".", ".."],
)
def test_row_image_references_must_be_base_filenames(tmp_path: Path, filename: str) -> None:
    row = ["APP-1", filename, "Brand", "Bourbon", "45", "750 mL"]
    parsed = parse_bytes(tmp_path, csv_package([row]), filename="paths.csv")

    assert PreflightIssueCode.INVALID_IMAGE_FILENAME in row_issue_codes(parsed)
    assert parsed.rows[0].normalized_label_image_filename is None
    assert not is_base_filename(filename)


def test_filename_matching_uses_trimmed_nfc_casefold_keys() -> None:
    assert normalize_filename("  CAFÉ.PNG  ") == normalize_filename("cafe\u0301.png")


@pytest.mark.parametrize(
    ("abv", "net_contents", "expected_codes"),
    [
        ("1e2", "750 mL", {PreflightIssueCode.INVALID_ABV}),
        ("101", "750 mL", {PreflightIssueCode.INVALID_ABV}),
        ("45", "750 oz", {PreflightIssueCode.INVALID_NET_CONTENTS}),
        ("45", "0 mL", {PreflightIssueCode.INVALID_NET_CONTENTS}),
    ],
)
def test_expected_numeric_fields_are_conservative(
    tmp_path: Path,
    abv: str,
    net_contents: str,
    expected_codes: set[PreflightIssueCode],
) -> None:
    row = ["APP-1", "label.png", "Brand", "Bourbon", abv, net_contents]
    parsed = parse_bytes(tmp_path, csv_package([row]), filename="values.csv")

    assert expected_codes <= set(row_issue_codes(parsed))
    assert parsed.rows[0].normalized_expected is None


def test_overlong_cells_are_bounded_and_rejected(tmp_path: Path) -> None:
    brand = "B" * 501
    row = ["APP-1", "label.png", brand, "Bourbon", "45", "750 mL"]
    parsed = parse_bytes(tmp_path, csv_package([row]), filename="long.csv")

    result = parsed.rows[0]
    assert len(result.expected_input.brand_name) == 200
    assert {PreflightIssueCode.CELL_TOO_LONG, PreflightIssueCode.INVALID_BRAND} <= set(
        row_issue_codes(parsed)
    )


def test_xlsx_formulas_are_rejected_without_evaluation(tmp_path: Path) -> None:
    content = xlsx_package([["APP-1", "label.png", "=1+1", "Bourbon", "45", "750 mL"]])
    parsed = parse_bytes(tmp_path, content, filename="formula.xlsx")

    issue = next(
        issue
        for issue in parsed.rows[0].issues
        if issue.code == PreflightIssueCode.FORMULA_NOT_ALLOWED
    )
    assert issue.field == BatchField.EXPECTED_BRAND
    assert parsed.rows[0].normalized_expected is None


def test_xlsx_requires_the_batch_worksheet(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Applications"
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_bytes(tmp_path, output.getvalue(), filename="wrong-sheet.xlsx")

    assert issue_codes(parsed) == [PreflightIssueCode.MISSING_BATCH_WORKSHEET]


def test_xlsx_rejects_macro_and_external_link_parts(tmp_path: Path) -> None:
    source = xlsx_package([["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]])

    def with_part(name: str) -> bytes:
        path = tmp_path / "modified.xlsx"
        path.write_bytes(source)
        with ZipFile(path, "a", compression=ZIP_DEFLATED) as archive:
            archive.writestr(name, b"not executable")
        return path.read_bytes()

    macro = parse_bytes(tmp_path, with_part("xl/vbaProject.bin"), filename="macro.xlsx")
    external = parse_bytes(
        tmp_path,
        with_part("xl/externalLinks/externalLink1.xml"),
        filename="external.xlsx",
    )

    assert issue_codes(macro) == [PreflightIssueCode.MACRO_ENABLED_WORKBOOK]
    assert issue_codes(external) == [PreflightIssueCode.EXTERNAL_LINK_NOT_ALLOWED]


def test_xlsx_archive_expansion_is_bounded_before_openpyxl_loads_it(tmp_path: Path) -> None:
    source = xlsx_package([["APP-1", "label.png", "Brand", "Bourbon", "45", "750 mL"]])
    path = tmp_path / "expanded.xlsx"
    path.write_bytes(source)
    with ZipFile(path, "a", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/oversized.bin", b"0" * (10 * 1024 * 1024))

    parsed = parse_spreadsheet(path, filename="expanded.xlsx")

    assert issue_codes(parsed) == [PreflightIssueCode.WORKBOOK_EXPANSION_LIMIT_EXCEEDED]


def test_xlsx_rejects_compound_encrypted_or_binary_content(tmp_path: Path) -> None:
    parsed = parse_bytes(
        tmp_path,
        bytes.fromhex("D0CF11E0A1B11AE1") + b"encrypted",
        filename="encrypted.xlsx",
    )

    assert issue_codes(parsed) == [PreflightIssueCode.ENCRYPTED_WORKBOOK]


def test_xlsx_source_row_span_is_bounded(tmp_path: Path) -> None:
    workbook = load_workbook(BytesIO(generate_xlsx_template()))
    workbook["Batch"]["A251"] = "APP-250"
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_bytes(tmp_path, output.getvalue(), filename="sparse.xlsx")

    assert issue_codes(parsed) == [PreflightIssueCode.SOURCE_ROW_LIMIT_EXCEEDED]
