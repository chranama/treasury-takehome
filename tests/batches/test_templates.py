import csv
from io import BytesIO, StringIO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.batches import (
    BATCH_TEMPLATE_HEADER_ALIASES,
    BATCH_TEMPLATE_HEADERS,
    BATCH_WORKSHEET_NAME,
    INSTRUCTIONS_WORKSHEET_NAME,
    generate_csv_template,
    generate_xlsx_template,
)


def test_csv_template_is_deterministic_and_round_trips_exact_headers() -> None:
    first = generate_csv_template()
    second = generate_csv_template()

    assert first == second
    rows = list(csv.reader(StringIO(first.decode("utf-8"))))
    assert rows == [list(BATCH_TEMPLATE_HEADERS)]


def test_xlsx_template_is_deterministic_and_round_trips_exact_headers() -> None:
    first = generate_xlsx_template()
    second = generate_xlsx_template()

    assert first == second
    workbook = load_workbook(BytesIO(first), read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == [BATCH_WORKSHEET_NAME, INSTRUCTIONS_WORKSHEET_NAME]
        rows = list(workbook[BATCH_WORKSHEET_NAME].iter_rows(values_only=True))
        assert rows == [BATCH_TEMPLATE_HEADERS]
        assert workbook[INSTRUCTIONS_WORKSHEET_NAME]["A2"].value.startswith("Use the Batch sheet")
    finally:
        workbook.close()

    with ZipFile(BytesIO(first)) as archive:
        core_properties = archive.read("docProps/core.xml")
    assert b"2000-01-01T00:00:00Z</dcterms:modified>" in core_properties


def test_templates_define_the_same_planned_parser_header_contract() -> None:
    csv_headers = next(csv.reader(StringIO(generate_csv_template().decode("utf-8"))))
    workbook = load_workbook(BytesIO(generate_xlsx_template()), read_only=True)
    try:
        xlsx_headers = next(
            workbook[BATCH_WORKSHEET_NAME].iter_rows(min_row=1, max_row=1, values_only=True)
        )
    finally:
        workbook.close()

    assert tuple(csv_headers) == xlsx_headers == BATCH_TEMPLATE_HEADERS
    assert BATCH_TEMPLATE_HEADER_ALIASES == {}
