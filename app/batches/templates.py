"""Deterministic blank batch-template generation."""

import csv
from datetime import datetime
from io import BytesIO, StringIO
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BATCH_TEMPLATE_HEADERS = (
    "Application ID",
    "Label Image Filename",
    "Expected Brand",
    "Expected Class/Type",
    "Expected ABV",
    "Expected Net Contents",
)
BATCH_TEMPLATE_HEADER_ALIASES: dict[str, str] = {}

BATCH_WORKSHEET_NAME = "Batch"
INSTRUCTIONS_WORKSHEET_NAME = "Instructions"

_INSTRUCTIONS = (
    "Use the Batch sheet and keep its column names unchanged.",
    "Enter one application per row, with no more than 25 rows.",
    "Use each image's base filename, including its extension.",
    "Enter ABV as a percentage such as 45 or 45%.",
    "Enter net contents as a metric quantity such as 750 mL or 0.75 L.",
    "Use only synthetic or otherwise non-sensitive data.",
)
_FIXED_TIMESTAMP = datetime(2000, 1, 1, 0, 0, 0)
_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)


def generate_csv_template() -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(BATCH_TEMPLATE_HEADERS)
    return output.getvalue().encode("utf-8")


def generate_xlsx_template() -> bytes:
    workbook = Workbook()
    workbook.properties.creator = "Label Review"
    workbook.properties.lastModifiedBy = "Label Review"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP

    batch = workbook.active
    batch.title = BATCH_WORKSHEET_NAME
    batch.append(BATCH_TEMPLATE_HEADERS)
    batch.freeze_panes = "A2"
    batch.auto_filter.ref = "A1:F1"
    for cell in batch[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for column, width in zip("ABCDEF", (20, 28, 28, 38, 18, 24), strict=True):
        batch.column_dimensions[column].width = width

    instructions = workbook.create_sheet(INSTRUCTIONS_WORKSHEET_NAME)
    instructions.append(("Batch template instructions",))
    instructions["A1"].font = Font(bold=True)
    for instruction in _INSTRUCTIONS:
        instructions.append((instruction,))
    instructions.column_dimensions["A"].width = 82

    raw = BytesIO()
    workbook.save(raw)
    workbook.close()
    return _normalize_xlsx_archive(raw.getvalue())


def _normalize_xlsx_archive(raw: bytes) -> bytes:
    """Remove ZIP timestamps so equal templates have equal bytes."""

    normalized = BytesIO()
    with (
        ZipFile(BytesIO(raw), "r") as source,
        ZipFile(
            normalized,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
        ) as destination,
    ):
        for name in sorted(source.namelist()):
            source_info = source.getinfo(name)
            target_info = ZipInfo(filename=name, date_time=_ZIP_TIMESTAMP)
            target_info.compress_type = ZIP_DEFLATED
            target_info.external_attr = source_info.external_attr
            target_info.create_system = 0
            destination.writestr(target_info, source.read(name))
    return normalized.getvalue()
