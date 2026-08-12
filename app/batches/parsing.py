"""Bounded, provider-neutral CSV and XLSX row parsing for P1 preflight."""

import csv
import math
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from io import StringIO
from pathlib import Path, PurePosixPath, PureWindowsPath
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException

from app.batches.contracts import (
    BatchExpectedInput,
    BatchField,
    PreflightIssue,
    PreflightIssueCode,
    PreflightIssueScope,
    PreflightIssueSeverity,
)
from app.batches.limits import (
    MAX_ABV_CELL_CHARACTERS,
    MAX_APPLICATION_ID_CHARACTERS,
    MAX_BATCH_CASES,
    MAX_EXPECTED_TEXT_CHARACTERS,
    MAX_IMAGE_FILENAME_CHARACTERS,
    MAX_NET_CONTENTS_CELL_CHARACTERS,
    MAX_SPREADSHEET_CELL_CHARACTERS,
    MAX_SPREADSHEET_SOURCE_ROWS,
    MAX_XLSX_ARCHIVE_ENTRIES,
    MAX_XLSX_UNCOMPRESSED_BYTES,
)
from app.batches.templates import BATCH_TEMPLATE_HEADERS, BATCH_WORKSHEET_NAME
from app.comparison import ExpectedNetContents, ExpectedReview, NetContentsUnit

_COMPOUND_FILE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
_ABV_PATTERN = re.compile(r"^(?P<value>(?:\d+(?:\.\d+)?|\.\d+))\s*%?$")
_NET_CONTENTS_PATTERN = re.compile(
    r"^(?P<value>(?:\d+(?:\.\d+)?|\.\d+))\s*(?P<unit>mL|L)$",
    re.IGNORECASE,
)

_HEADER_FIELDS = dict(zip(BATCH_TEMPLATE_HEADERS, BatchField, strict=True))
_EXPECTED_FIELDS = frozenset(
    {
        BatchField.EXPECTED_BRAND,
        BatchField.EXPECTED_CLASS_TYPE,
        BatchField.EXPECTED_ABV,
        BatchField.EXPECTED_NET_CONTENTS,
    }
)


class SpreadsheetKind(StrEnum):
    CSV = "csv"
    XLSX = "xlsx"
    UNSUPPORTED = "unsupported"


@dataclass(frozen=True, slots=True)
class ParsedSpreadsheetRow:
    row_number: int
    application_id: str
    normalized_application_id: str | None
    label_image_filename: str
    normalized_label_image_filename: str | None
    expected_input: BatchExpectedInput
    normalized_expected: ExpectedReview | None
    issues: tuple[PreflightIssue, ...]

    @property
    def has_errors(self) -> bool:
        return any(issue.severity == PreflightIssueSeverity.ERROR for issue in self.issues)


@dataclass(frozen=True, slots=True)
class ParsedSpreadsheet:
    kind: SpreadsheetKind
    rows: tuple[ParsedSpreadsheetRow, ...]
    issues: tuple[PreflightIssue, ...]

    @property
    def has_errors(self) -> bool:
        return any(issue.severity == PreflightIssueSeverity.ERROR for issue in self.issues)


@dataclass(frozen=True, slots=True)
class _RawCell:
    text: str
    formula: bool = False


@dataclass(slots=True)
class _MutableRow:
    row_number: int
    application_id: str
    normalized_application_id: str | None
    label_image_filename: str
    normalized_label_image_filename: str | None
    expected_input: BatchExpectedInput
    normalized_expected: ExpectedReview | None
    issues: list[PreflightIssue] = field(default_factory=list)

    def freeze(self) -> ParsedSpreadsheetRow:
        return ParsedSpreadsheetRow(
            row_number=self.row_number,
            application_id=self.application_id,
            normalized_application_id=self.normalized_application_id,
            label_image_filename=self.label_image_filename,
            normalized_label_image_filename=self.normalized_label_image_filename,
            expected_input=self.expected_input,
            normalized_expected=self.normalized_expected,
            issues=tuple(self.issues),
        )


def parse_spreadsheet(path: Path, *, filename: str) -> ParsedSpreadsheet:
    """Parse a previously size-bounded spreadsheet without evaluating active content."""

    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        return _parse_csv(path)
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    return ParsedSpreadsheet(
        kind=SpreadsheetKind.UNSUPPORTED,
        rows=(),
        issues=(_batch_issue(PreflightIssueCode.UNSUPPORTED_SPREADSHEET),),
    )


def normalize_match_text(value: str) -> str:
    return unicodedata.normalize("NFC", value.strip()).casefold()


def normalize_filename(value: str) -> str | None:
    trimmed = unicodedata.normalize("NFC", value.strip())
    if not is_base_filename(trimmed):
        return None
    return trimmed.casefold()


def is_base_filename(value: str) -> bool:
    if not value or value in {".", ".."} or "\x00" in value:
        return False
    posix = PurePosixPath(value)
    windows = PureWindowsPath(value)
    return (
        posix.name == value
        and windows.name == value
        and not windows.drive
        and "/" not in value
        and "\\" not in value
    )


def _parse_csv(path: Path) -> ParsedSpreadsheet:
    try:
        raw = path.read_bytes()
        if b"\x00" in raw:
            return ParsedSpreadsheet(
                kind=SpreadsheetKind.CSV,
                rows=(),
                issues=(_batch_issue(PreflightIssueCode.NUL_BYTE_NOT_ALLOWED),),
            )
        text = raw.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError:
        return ParsedSpreadsheet(
            kind=SpreadsheetKind.CSV,
            rows=(),
            issues=(_batch_issue(PreflightIssueCode.INVALID_CSV_ENCODING),),
        )
    except OSError:
        return _malformed(SpreadsheetKind.CSV)

    try:
        reader = csv.reader(StringIO(text, newline=""), strict=True)
        records: list[tuple[int, list[_RawCell]]] = []
        for source_row, values in enumerate(reader, start=1):
            if source_row > MAX_SPREADSHEET_SOURCE_ROWS:
                return ParsedSpreadsheet(
                    kind=SpreadsheetKind.CSV,
                    rows=(),
                    issues=(_batch_issue(PreflightIssueCode.SOURCE_ROW_LIMIT_EXCEEDED),),
                )
            records.append((source_row, [_RawCell(value) for value in values]))
    except csv.Error:
        return _malformed(SpreadsheetKind.CSV)

    return _parse_records(SpreadsheetKind.CSV, records)


def _parse_xlsx(path: Path) -> ParsedSpreadsheet:
    try:
        if path.read_bytes()[:8] == _COMPOUND_FILE_SIGNATURE:
            return ParsedSpreadsheet(
                kind=SpreadsheetKind.XLSX,
                rows=(),
                issues=(_batch_issue(PreflightIssueCode.ENCRYPTED_WORKBOOK),),
            )
        with ZipFile(path) as archive:
            entries = archive.infolist()
            if any(entry.flag_bits & 0x1 for entry in entries):
                return ParsedSpreadsheet(
                    kind=SpreadsheetKind.XLSX,
                    rows=(),
                    issues=(_batch_issue(PreflightIssueCode.ENCRYPTED_WORKBOOK),),
                )
            if (
                len(entries) > MAX_XLSX_ARCHIVE_ENTRIES
                or sum(entry.file_size for entry in entries) > MAX_XLSX_UNCOMPRESSED_BYTES
            ):
                return ParsedSpreadsheet(
                    kind=SpreadsheetKind.XLSX,
                    rows=(),
                    issues=(_batch_issue(PreflightIssueCode.WORKBOOK_EXPANSION_LIMIT_EXCEEDED),),
                )
            names = {name.casefold() for name in archive.namelist()}
            if "xl/vbaproject.bin" in names:
                return ParsedSpreadsheet(
                    kind=SpreadsheetKind.XLSX,
                    rows=(),
                    issues=(_batch_issue(PreflightIssueCode.MACRO_ENABLED_WORKBOOK),),
                )
            if any(name.startswith("xl/externallinks/") for name in names):
                return ParsedSpreadsheet(
                    kind=SpreadsheetKind.XLSX,
                    rows=(),
                    issues=(_batch_issue(PreflightIssueCode.EXTERNAL_LINK_NOT_ALLOWED),),
                )
    except (BadZipFile, OSError):
        return _malformed(SpreadsheetKind.XLSX)

    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError):
        return _malformed(SpreadsheetKind.XLSX)

    try:
        if BATCH_WORKSHEET_NAME not in workbook.sheetnames:
            return ParsedSpreadsheet(
                kind=SpreadsheetKind.XLSX,
                rows=(),
                issues=(_batch_issue(PreflightIssueCode.MISSING_BATCH_WORKSHEET),),
            )
        worksheet = workbook[BATCH_WORKSHEET_NAME]
        if worksheet.max_row > MAX_SPREADSHEET_SOURCE_ROWS:
            return ParsedSpreadsheet(
                kind=SpreadsheetKind.XLSX,
                rows=(),
                issues=(_batch_issue(PreflightIssueCode.SOURCE_ROW_LIMIT_EXCEEDED),),
            )
        if worksheet.max_column > len(BATCH_TEMPLATE_HEADERS):
            return ParsedSpreadsheet(
                kind=SpreadsheetKind.XLSX,
                rows=(),
                issues=(_batch_issue(PreflightIssueCode.UNEXPECTED_COLUMN),),
            )

        records = [
            (row_number, [_xlsx_cell(cell) for cell in row])
            for row_number, row in enumerate(
                worksheet.iter_rows(max_col=len(BATCH_TEMPLATE_HEADERS)),
                start=1,
            )
        ]
        return _parse_records(SpreadsheetKind.XLSX, records)
    except (OSError, TypeError, ValueError):
        return _malformed(SpreadsheetKind.XLSX)
    finally:
        workbook.close()


def _xlsx_cell(cell: Cell) -> _RawCell:
    return _RawCell(text=_display_cell_value(cell.value), formula=cell.data_type == "f")


def _display_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        return format(value, ".15g")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _parse_records(
    kind: SpreadsheetKind,
    records: list[tuple[int, list[_RawCell]]],
) -> ParsedSpreadsheet:
    if not records:
        return ParsedSpreadsheet(
            kind=kind,
            rows=(),
            issues=(_batch_issue(PreflightIssueCode.EMPTY_BATCH),),
        )

    _, header_cells = records[0]
    header_issues = _validate_headers(header_cells)
    if header_issues:
        return ParsedSpreadsheet(kind=kind, rows=(), issues=tuple(header_issues))

    batch_issues: list[PreflightIssue] = []
    rows: list[_MutableRow] = []
    for row_number, cells in records[1:]:
        if _row_is_blank(cells):
            continue
        if len(rows) == MAX_BATCH_CASES:
            batch_issues.append(_batch_issue(PreflightIssueCode.TOO_MANY_CASES))
            break
        rows.append(_parse_data_row(row_number, cells))

    if not rows and not batch_issues:
        batch_issues.append(_batch_issue(PreflightIssueCode.EMPTY_BATCH))

    _mark_duplicate_application_ids(rows)
    _mark_duplicate_image_references(rows)
    return ParsedSpreadsheet(
        kind=kind,
        rows=tuple(row.freeze() for row in rows),
        issues=tuple(batch_issues),
    )


def _validate_headers(cells: list[_RawCell]) -> list[PreflightIssue]:
    issues: list[PreflightIssue] = []
    if any(cell.formula for cell in cells):
        issues.append(_batch_issue(PreflightIssueCode.FORMULA_NOT_ALLOWED))
        return issues

    headers = [cell.text for cell in cells]
    if any(len(header) > MAX_SPREADSHEET_CELL_CHARACTERS for header in headers):
        issues.append(_batch_issue(PreflightIssueCode.CELL_TOO_LONG))
        return issues

    actual = tuple(headers)
    if actual == BATCH_TEMPLATE_HEADERS:
        return issues

    for header, field_name in _HEADER_FIELDS.items():
        count = actual.count(header)
        if count == 0:
            issues.append(
                PreflightIssue(
                    code=PreflightIssueCode.MISSING_REQUIRED_COLUMN,
                    scope=PreflightIssueScope.BATCH,
                    field=field_name,
                )
            )
        elif count > 1:
            issues.append(
                PreflightIssue(
                    code=PreflightIssueCode.DUPLICATE_COLUMN,
                    scope=PreflightIssueScope.BATCH,
                    field=field_name,
                )
            )
    if any(header not in _HEADER_FIELDS for header in actual):
        issues.append(_batch_issue(PreflightIssueCode.UNEXPECTED_COLUMN))
    if set(actual) == set(BATCH_TEMPLATE_HEADERS) and len(actual) == len(BATCH_TEMPLATE_HEADERS):
        issues.append(_batch_issue(PreflightIssueCode.INVALID_COLUMN_ORDER))
    return issues


def _row_is_blank(cells: list[_RawCell]) -> bool:
    return not any(cell.text.strip() or cell.formula for cell in cells)


def _parse_data_row(row_number: int, cells: list[_RawCell]) -> _MutableRow:
    padded = cells[: len(BATCH_TEMPLATE_HEADERS)] + [_RawCell("")] * max(
        0, len(BATCH_TEMPLATE_HEADERS) - len(cells)
    )
    issues: list[PreflightIssue] = []
    if len(cells) > len(BATCH_TEMPLATE_HEADERS) and any(
        cell.text.strip() or cell.formula for cell in cells[len(BATCH_TEMPLATE_HEADERS) :]
    ):
        issues.append(
            PreflightIssue(
                code=PreflightIssueCode.UNEXPECTED_COLUMN,
                scope=PreflightIssueScope.ROW,
                row_number=row_number,
            )
        )

    for field_name, cell in zip(BatchField, padded, strict=True):
        if cell.formula:
            issues.append(
                _row_issue(PreflightIssueCode.FORMULA_NOT_ALLOWED, row_number, field_name)
            )

    values = [_normalize_cell(cell.text) for cell in padded]
    application_id = _bound_value(
        values[0],
        MAX_APPLICATION_ID_CHARACTERS,
        row_number,
        BatchField.APPLICATION_ID,
        issues,
    )
    label_filename = _bound_value(
        values[1],
        MAX_IMAGE_FILENAME_CHARACTERS,
        row_number,
        BatchField.LABEL_IMAGE_FILENAME,
        issues,
    )
    brand = _bound_expected_value(
        values[2],
        MAX_EXPECTED_TEXT_CHARACTERS,
        PreflightIssueCode.INVALID_BRAND,
        row_number,
        BatchField.EXPECTED_BRAND,
        issues,
    )
    class_type = _bound_expected_value(
        values[3],
        MAX_EXPECTED_TEXT_CHARACTERS,
        PreflightIssueCode.INVALID_CLASS_TYPE,
        row_number,
        BatchField.EXPECTED_CLASS_TYPE,
        issues,
    )
    abv_text = _bound_expected_value(
        values[4],
        MAX_ABV_CELL_CHARACTERS,
        PreflightIssueCode.INVALID_ABV,
        row_number,
        BatchField.EXPECTED_ABV,
        issues,
    )
    net_contents_text = _bound_expected_value(
        values[5],
        MAX_NET_CONTENTS_CELL_CHARACTERS,
        PreflightIssueCode.INVALID_NET_CONTENTS,
        row_number,
        BatchField.EXPECTED_NET_CONTENTS,
        issues,
    )

    normalized_application_id = normalize_match_text(application_id) if application_id else None
    if not application_id:
        issues.append(
            _row_issue(
                PreflightIssueCode.MISSING_APPLICATION_ID,
                row_number,
                BatchField.APPLICATION_ID,
            )
        )

    normalized_label_filename = normalize_filename(label_filename)
    if not label_filename:
        issues.append(
            _row_issue(
                PreflightIssueCode.MISSING_IMAGE_FILENAME,
                row_number,
                BatchField.LABEL_IMAGE_FILENAME,
            )
        )
    elif normalized_label_filename is None:
        issues.append(
            _row_issue(
                PreflightIssueCode.INVALID_IMAGE_FILENAME,
                row_number,
                BatchField.LABEL_IMAGE_FILENAME,
            )
        )

    abv = _parse_expected_abv(abv_text)
    if not brand:
        issues.append(
            _row_issue(
                PreflightIssueCode.INVALID_BRAND,
                row_number,
                BatchField.EXPECTED_BRAND,
            )
        )
    if not class_type:
        issues.append(
            _row_issue(
                PreflightIssueCode.INVALID_CLASS_TYPE,
                row_number,
                BatchField.EXPECTED_CLASS_TYPE,
            )
        )
    if abv is None:
        issues.append(
            _row_issue(
                PreflightIssueCode.INVALID_ABV,
                row_number,
                BatchField.EXPECTED_ABV,
            )
        )
    net_contents = _parse_expected_net_contents(net_contents_text)
    if net_contents is None:
        issues.append(
            _row_issue(
                PreflightIssueCode.INVALID_NET_CONTENTS,
                row_number,
                BatchField.EXPECTED_NET_CONTENTS,
            )
        )

    expected_input = BatchExpectedInput(
        brand_name=brand,
        class_type=class_type,
        expected_abv=abv_text,
        expected_net_contents=net_contents_text,
    )
    expected_has_errors = any(
        issue.severity == PreflightIssueSeverity.ERROR and issue.field in _EXPECTED_FIELDS
        for issue in issues
    )
    normalized_expected = None
    if not expected_has_errors and abv is not None and net_contents is not None:
        normalized_expected = ExpectedReview(
            brand_name=brand,
            class_type=class_type,
            abv=abv,
            net_contents=net_contents,
        )

    return _MutableRow(
        row_number=row_number,
        application_id=application_id,
        normalized_application_id=normalized_application_id,
        label_image_filename=label_filename,
        normalized_label_image_filename=normalized_label_filename,
        expected_input=expected_input,
        normalized_expected=normalized_expected,
        issues=_deduplicate_issues(issues),
    )


def _normalize_cell(value: str) -> str:
    return unicodedata.normalize("NFC", value.strip())


def _bound_value(
    value: str,
    maximum: int,
    row_number: int,
    field_name: BatchField,
    issues: list[PreflightIssue],
) -> str:
    if len(value) > maximum:
        issues.append(_row_issue(PreflightIssueCode.CELL_TOO_LONG, row_number, field_name))
    return value[:maximum]


def _bound_expected_value(
    value: str,
    maximum: int,
    issue_code: PreflightIssueCode,
    row_number: int,
    field_name: BatchField,
    issues: list[PreflightIssue],
) -> str:
    bounded = _bound_value(value, maximum, row_number, field_name, issues)
    if len(value) > maximum:
        issues.append(_row_issue(issue_code, row_number, field_name))
    return bounded


def _parse_expected_abv(value: str) -> Decimal | None:
    match = _ABV_PATTERN.fullmatch(value)
    if match is None:
        return None
    try:
        parsed = Decimal(match.group("value"))
    except InvalidOperation:
        return None
    return parsed if Decimal(0) <= parsed <= Decimal(100) else None


def _parse_expected_net_contents(value: str) -> ExpectedNetContents | None:
    match = _NET_CONTENTS_PATTERN.fullmatch(value)
    if match is None:
        return None
    try:
        parsed = Decimal(match.group("value"))
    except InvalidOperation:
        return None
    if parsed <= 0:
        return None
    unit = (
        NetContentsUnit.MILLILITER
        if match.group("unit").casefold() == "ml"
        else NetContentsUnit.LITER
    )
    return ExpectedNetContents(value=parsed, unit=unit)


def _mark_duplicate_application_ids(rows: list[_MutableRow]) -> None:
    grouped: dict[str, list[_MutableRow]] = {}
    for row in rows:
        if row.normalized_application_id is not None:
            grouped.setdefault(row.normalized_application_id, []).append(row)
    for duplicates in grouped.values():
        if len(duplicates) < 2:
            continue
        for row in duplicates:
            row.issues.append(
                _row_issue(
                    PreflightIssueCode.DUPLICATE_APPLICATION_ID,
                    row.row_number,
                    BatchField.APPLICATION_ID,
                )
            )


def _mark_duplicate_image_references(rows: list[_MutableRow]) -> None:
    grouped: dict[str, list[_MutableRow]] = {}
    for row in rows:
        if row.normalized_label_image_filename is not None:
            grouped.setdefault(row.normalized_label_image_filename, []).append(row)
    for duplicates in grouped.values():
        if len(duplicates) < 2:
            continue
        for row in duplicates:
            row.issues.append(
                _row_issue(
                    PreflightIssueCode.DUPLICATE_IMAGE_FILENAME,
                    row.row_number,
                    BatchField.LABEL_IMAGE_FILENAME,
                )
            )


def _deduplicate_issues(issues: list[PreflightIssue]) -> list[PreflightIssue]:
    result: list[PreflightIssue] = []
    keys: set[tuple[PreflightIssueCode, PreflightIssueScope, int | None, BatchField | None]] = set()
    for issue in issues:
        key = (issue.code, issue.scope, issue.row_number, issue.field)
        if key not in keys:
            keys.add(key)
            result.append(issue)
    return result


def _batch_issue(code: PreflightIssueCode) -> PreflightIssue:
    return PreflightIssue(code=code, scope=PreflightIssueScope.BATCH)


def _row_issue(
    code: PreflightIssueCode,
    row_number: int,
    field_name: BatchField | None = None,
) -> PreflightIssue:
    return PreflightIssue(
        code=code,
        scope=PreflightIssueScope.ROW,
        row_number=row_number,
        field=field_name,
    )


def _malformed(kind: SpreadsheetKind) -> ParsedSpreadsheet:
    return ParsedSpreadsheet(
        kind=kind,
        rows=(),
        issues=(_batch_issue(PreflightIssueCode.MALFORMED_SPREADSHEET),),
    )
